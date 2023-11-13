import openpyxl
import datetime
from datetime import datetime
from datetime import date


def toSheet(output_sheet, currentRow, ubi_Whole, taxCode, lineItemAmt, dataTranCode, dataTranDate, dataTranLSN, ubi_valuesDict):
    cr = str(currentRow)
    output_sheet["E" + cr] = ubi_Whole
    output_sheet["O" + cr] = remove_mod(taxCode)
    output_sheet["M" + cr] = lineItemAmt
    output_sheet["A" + cr] = dataTranCode
    output_sheet["B" + cr] = dataTranDate
    output_sheet["C" + cr] = ubi_valuesDict[ubi_Whole][1]
    output_sheet["D" + cr] = dataTranLSN
    output_sheet["G" + cr] = ubi_valuesDict[ubi_Whole][2]
    output_sheet["H" + cr] = ubi_valuesDict[ubi_Whole][3]
    output_sheet["I" + cr] = ubi_valuesDict[ubi_Whole][4]
    output_sheet["J" + cr] = ubi_valuesDict[ubi_Whole][5]
    output_sheet["K" + cr] = ubi_valuesDict[ubi_Whole][6]
    output_sheet["L" + cr] = ubi_valuesDict[ubi_Whole][7]
    output_sheet["P" + cr] = ubi_valuesDict[ubi_Whole][0]


def taxCalc(lineTaxType, taxCodeAmt, lineItemAmt, proration, currentRow, output_sheet):
    if lineTaxType == "percent":
        output_sheet["N" + str(currentRow)] = taxCodeAmt * lineItemAmt
    elif lineTaxType == "fixed":
        output_sheet["N" + str(currentRow)] = taxCodeAmt * proration


def prorate(start_date, end_date):
    start = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")

    total_days = (end - start).days + 1
    month_days = end.day

    return total_days / month_days


def remove_mod(s):
    if '*' in s:
        return s[:s.index('*')]
    else:
        return s


excelName = input("Enter the excel file name(including the file extension): ")
# excelName = "BI-IN8802withFees.xlsx"
sheetName = input("Enter the sheet name(eg. Sheet1): ")
# sheetName = "Sheet1"
print("Loading Excel sheets into memory...")
try:
    ubi_input = openpyxl.load_workbook(filename=excelName, read_only=True)
    ubi_sheet = ubi_input[sheetName]
except Exception as e:
    print("Error with spreadsheet name.")
    input("Press any key to exit...")
try:
    ubi_output = openpyxl.load_workbook("TemplateFinal.xlsx")
    output_sheet = ubi_output["Sheet1"]
except Exception as e:
    print("TemplateFinal.xlsx is missing or sheet name has been altered from Sheet1")
    input("Press any key to exit...")
try:
    ubi_output_horizontal = openpyxl.load_workbook("TemplateFinal_Horizontal.xlsx")
    output_sheet_horizontal = ubi_output_horizontal["Sheet1"]
except Exception as e:
    print("TemplateFinal_Horizontal.xlsx is missing or sheet name has been altered from Sheet1")
    input("Press any key to exit...")
database = openpyxl.load_workbook("Database.xlsx")
tax_database = database["Sheet3"]
nsc_database = database["Sheet2"]
fed_database = database["Sheet4"]
ubi_line_count = 1
nsc_line_count = 1
tax_line_count = 1
nsc_dict = {}
taxCode_dict = {}
fed_dict = {}
ubi_list = []
missing_nsc = []
ubi_dateDict = {}
ubi_amountDict = {}
ubi_valuesDict = {}
currentRow = 1
currentRow_horizontal = 3
currentRead = 1
totalLineItemAmt = 0
currentDate = date.today()

# Output table vars
dataTranCode = "TAX"
dataTranDate = currentDate
dataTranLSN = ""

# write the ubi column to a list for faster referencing
print("Generating UBI List...")
for row in ubi_sheet.iter_rows(min_row=2, min_col=1, max_col=ubi_sheet.max_column):
    ubi = row[12].value
    date = row[50].value
    order_number = ubi[ubi.find('EIS'):ubi.find('EIS') + 3]
    for c in ubi[ubi.find('EIS') + 3:]:
        if not c.isdigit():
            break
        order_number += c
    start = ubi.find(order_number) + len(order_number)
    nsc = ubi[start:start + 8]
    start = ubi.find('_') + 1
    clin = ubi[start:start + 7]
    start = ubi.find(clin) + len(clin) + 1
    ubiFixed = ubi[start:]
    ubi_dateDict[ubiFixed] = date.strftime('%Y-%m-%d')
    ubi_amountDict[ubiFixed] = row[88].value
    if ubi not in ubi_list:
        ubi_list.append(ubi)

# info per ubi line
for row in ubi_sheet.iter_rows(min_row=2, min_col=1, max_col=ubi_sheet.max_column):
    ubi = row[12].value
    contractNo = row[2].value
    ogBillLineItemSN = row[9].value
    termNSC = row[26].value
    contractIN = row[48].value
    contractID = row[49].value
    contractID = contractID.strftime('%Y-%m-%d')
    billingP = row[50].value
    billingP = billingP.strftime('%Y-%m-%d')
    billingBD = row[51].value
    billingBD = billingBD.strftime('%Y-%m-%d')
    billingED = row[52].value
    billingED = billingED.strftime('%Y-%m-%d')
    lineItemAMT1 = row[88].value
    ubi_valuesDict[ubi] = (
        ogBillLineItemSN, contractNo, termNSC, contractIN, contractID, billingP, billingBD, billingED, lineItemAMT1)

# hash the nsc table to a dict for faster referencing
print("Generating NSC table...")
for row in nsc_database.iter_rows(min_row=2, min_col=1, max_col=nsc_database.max_column):
    nsc = row[0].value
    city = row[2].value
    state = row[30].value
    county = row[42].value
    nsc_dict[nsc] = (state, county, city)

# hash the tax table to a dict for faster referencing
print("Generating tax table...")
for row in tax_database.iter_rows(min_row=2, min_col=1, max_col=tax_database.max_column):
    taxAmount = 0
    tax_code = row[0].value
    cityTax = row[5].value
    if type(cityTax) != str:
        cityTax = "NA"
    cityTax = cityTax.upper()
    stateTax = row[3].value
    if type(stateTax) != str:
        stateTax = "NA"
    stateTax = stateTax.upper()
    countyTax = row[4].value
    if type(countyTax) != str:
        countyTax = "NA"
    countyTax = countyTax.upper()
    authTax = row[6].value
    modNumber = row[1].value
    if type(row[12].value) == float or type(row[12].value) == int:
        taxAmount = row[12].value
        taxType = "percent"
    if type(row[13].value) == float or type(row[13].value) == int:
        taxAmount = row[13].value
        taxType = "fixed"
    dateStart = row[18].value
    dateStart = datetime.fromisoformat(dateStart)
    dateStart = dateStart.strftime('%Y-%m-%d')
    dateEnd = row[19].value
    dateEnd = datetime.fromisoformat(dateEnd)
    dateEnd = dateEnd.strftime('%Y-%m-%d')
    serviceLv = row[7].value
    if stateTax not in taxCode_dict and stateTax is not None:
        taxCode_dict[stateTax] = {}
    if countyTax not in taxCode_dict[stateTax] and countyTax is not None:
        taxCode_dict[stateTax][countyTax] = {}
    if cityTax not in taxCode_dict[stateTax][countyTax] and cityTax is not None:
        taxCode_dict[stateTax][countyTax][cityTax] = {}
    if countyTax not in taxCode_dict and countyTax is None and cityTax is None:
        taxCode_dict[stateTax][tax_code + "*" + str(modNumber)] = (
            stateTax, countyTax, cityTax, authTax, modNumber, taxAmount, dateStart, dateEnd, serviceLv, taxType)
    elif cityTax not in taxCode_dict and cityTax is not None and cityTax is None:
        taxCode_dict[stateTax][countyTax][tax_code + "*" + str(modNumber)] = (
            stateTax, countyTax, cityTax, authTax, modNumber, taxAmount, dateStart, dateEnd, serviceLv, taxType)
    elif cityTax is not None and countyTax is not None and stateTax is not None:
        taxCode_dict[stateTax][countyTax][cityTax][tax_code + "*" + str(modNumber)] = (
            stateTax, countyTax, cityTax, authTax, modNumber, taxAmount, dateStart, dateEnd, serviceLv, taxType)

# hash the fed tax table to a dict for faster referencing
print("Generating Federal Tax Table....")
for row in fed_database.iter_rows(min_row=2, min_col=1, max_col=fed_database.max_column):
    fed_clin = row[0].value
    federalTax1 = row[2].value
    federalTax2 = row[3].value
    federalTax3 = row[4].value
    federalTax4 = row[5].value
    fed_dict[fed_clin] = (federalTax1, federalTax2, federalTax3, federalTax4)

print("Validating all NSC")
for ubi_line in ubi_list:
    ubi_Whole = ubi_line
    order_number = ubi_Whole[ubi_Whole.find('EIS'):ubi_Whole.find('EIS') + 3]
    for c in ubi_Whole[ubi_Whole.find('EIS') + 3:]:
        if not c.isdigit():
            break
        order_number += c
    start = ubi_Whole.find(order_number) + len(order_number)
    nsc = ubi_Whole[start:start + 8]
    if nsc not in nsc_dict:
        if nsc not in missing_nsc:
            missing_nsc.append(nsc)
if missing_nsc:
    print("Error missing NSCs: " + str(missing_nsc))
    print("Please add these NSC entries to the database and re-run.")
    input()
    exit()


try:
    # main loop
    for ubi_line in ubi_list:
        # initialize tax code values
        valid = True
        taxCodeCO = ""
        taxCodeCOAmt = 0
        taxCodeCI = ""
        taxCodeCIAmt = 0
        taxCodeST = ""
        taxCodeDC = ""
        taxCodeSTAmt = 0
        taxCodeFED = ""
        highestModCI = 0
        highestModCO = 0
        highestModCO2 = 0
        highestModST = 0
        lineTax = 0
        exempt = True
        skip = False
        # Get UBI line and break up the UBI string into its component parts
        ubi_Whole = ubi_line
        proration = prorate(ubi_valuesDict[ubi_Whole][6], ubi_valuesDict[ubi_Whole][7])
        # make this into a method later, it's used to break down the whole ubi into its components
        order_number = ubi_Whole[ubi_Whole.find('EIS'):ubi_Whole.find('EIS') + 3]
        for c in ubi_Whole[ubi_Whole.find('EIS') + 3:]:
            if not c.isdigit():
                break
            order_number += c
        start = ubi_Whole.find(order_number) + len(order_number)
        nsc = ubi_Whole[start:start + 8]
        start = ubi_Whole.find('_') + 1
        clin = ubi_Whole[start:start + 7]
        start = ubi_Whole.find(clin) + len(clin) + 1
        ubi = ubi_Whole[start:]
        lineItemAmt = ubi_amountDict[ubi]
        totalLineItemAmt = lineItemAmt + totalLineItemAmt
        #  Toggle this to T/F
        if "IP" in clin or "EN" in clin or "TF" in clin or "NS" in clin or "VI" in clin or "WL" in clin or "MN" in clin:
            skip = False
        # nsc look up loop
        if nsc in nsc_dict:
            state, county, city = nsc_dict[nsc]
        if type(city) != str:
            city = "NA"
        stateTax = row[3].value
        if type(state) != str:
            state = "NA"
        countyTax = row[4].value
        if type(county) != str:
            county = "NA"

        # obtain state tax code
        if state in taxCode_dict:
            if "NA" in taxCode_dict[state]:
                county1 = "NA"
                if "NA" in taxCode_dict[state][county1]:
                    for key, values in taxCode_dict[state]["NA"]["NA"].items():
                        if values[6] <= ubi_dateDict[ubi] <= values[7]:
                            if values[3] == "ST":
                                if clin == "VS12110" or clin == "VS11310" or clin == "VS11210":
                                    if values[9] == "fixed":
                                        exempt = False
                                if exempt == False or values[9] == "percent":
                                    if values[8] != "E":
                                        if values[4] > highestModST:
                                            highestModST = values[4]
                                            taxCodeST = key
                                            taxCodeSTAmt = values[5]
                                            taxCodeSTAmt = taxCode_dict[state]["NA"]["NA"][key][5]
                                            lineTaxTypeST = taxCode_dict[state]["NA"]["NA"][key][9]
        # obtain county tax code
        if state in taxCode_dict:
            if county in taxCode_dict[state]:
                if city not in taxCode_dict[state][county]:
                    city1 = "NA"
                    if city1 in taxCode_dict[state][county]:
                        for key, values in taxCode_dict[state][county][city1].items():
                            if values[6] <= ubi_dateDict[ubi] <= values[7]:
                                if values[3] == "CO":
                                    if clin == "VS12110" or clin == "VS11310" or clin == "VS11210":
                                        if values[9] == "fixed":
                                            exempt = False
                                    if exempt == False or values[9] == "percent":
                                        if clin[:2] == "EN" and values[8] == "E":
                                            valid = True
                                        if clin[:2] == "L" and values[8] == "L":
                                            valid = True
                                        if values[8] == "B" or values[8] == "S":
                                            valid = True
                                        if values[4] > highestModCO and valid == True:
                                            highestModCO = values[4]
                                            taxCodeCO = key
                                            taxCodeCOAmt = values[5]
                                            taxCodeCOAmt = taxCode_dict[state][county][city1][key][5]
                                            lineTaxTypeCO = taxCode_dict[state][county][city1][key][9]

                            valid = True
                else:
                    if city in taxCode_dict[state][county]:
                        for key, values in taxCode_dict[state][county][city].items():
                            if values[6] <= ubi_dateDict[ubi] <= values[7]:
                                if values[3] == "CO":
                                    if clin == "VS12110" or clin == "VS11310" or clin == "VS11210":
                                        if values[9] == "fixed":
                                            exempt = False
                                    if exempt == False or values[9] == "percent":
                                        if clin[:2] == "EN" and values[8] == "E":
                                            valid = True
                                        if values[8] == "L":
                                            valid = True
                                        if values[8] == "B" or values[8] == "S":
                                            valid = True
                                        if values[4] > highestModCO and valid == True:
                                            highestModCO = values[4]
                                            taxCodeCO = key
                                            taxCodeCOAmt = values[5]
                                            taxCodeCOAmt = taxCode_dict[state][county][city][key][5]
                                            lineTaxTypeCO = taxCode_dict[state][county][city][key][9]
                            valid = True
                if taxCodeCO == "":
                    city1 = "NA"
                    if city1 in taxCode_dict[state][county]:
                        for key, values in taxCode_dict[state][county][city1].items():
                            if values[6] <= ubi_dateDict[ubi] <= values[7]:
                                if values[3] == "CO":
                                    if clin == "VS12110" or clin == "VS11310" or clin == "VS11210":
                                        if values[9] == "fixed":
                                            exempt = False
                                    if exempt == False or values[9] == "percent":
                                        if clin[:2] == "EN" and values[8] == "E":
                                            valid = True
                                        if clin[:2] == "L" and values[8] == "L":
                                            valid = True
                                        if values[8] == "B" or values[8] == "S":
                                            valid = True
                                        if values[4] > highestModCO and valid == True:
                                            highestModCO = values[4]
                                            taxCodeCO = key
                                            taxCodeCOAmt = values[5]
                                            taxCodeCOAmt = taxCode_dict[state][county][city1][key][5]
                                            lineTaxTypeCO = taxCode_dict[state][county][city1][key][9]

        # obtain city tax code with no county
        if state in taxCode_dict:
            county1 = "NA"
            if county1 in taxCode_dict[state]:
                if city in taxCode_dict[state][county1]:
                    for key, values in taxCode_dict[state][county1][city].items():
                        if values[6] <= ubi_dateDict[ubi] <= values[7]:
                            if values[3] == "CI":
                                if clin == "VS12110" or clin == "VS11310" or clin == "VS11210":
                                    if values[9] == "fixed":
                                        exempt = False
                                if exempt == False or values[9] == "percent":
                                    if clin[:2] == "EN" and values[8] == "E":
                                        valid = True
                                    if values[8] == "L":
                                        valid = True
                                    if values[8] == "B":
                                        valid = True
                                    if values[8] == "A":
                                        valid = True
                                    if values[8] == "S":
                                        valid = True
                                    if taxCodeCI == "" or values[8] != "B":
                                        if values[8] == "S" or values[8] == "A":
                                            valid = True
                                    if values[4] > highestModCI and valid == True:  # only take the highest tax mod
                                        highestModCI = values[4]
                                        taxCodeCI = key
                                        taxCodeCIAmt = values[5]
                                        taxCodeCIAmt = taxCode_dict[state][county1][city][key][5]
                                        lineTaxTypeCI = taxCode_dict[state][county1][city][key][9]
                        valid = True
        # obtain city tax code
        if state in taxCode_dict:
            if county in taxCode_dict[state]:
                if city in taxCode_dict[state][county]:
                    for key, values in taxCode_dict[state][county][city].items():
                        if values[6] <= ubi_dateDict[ubi] <= values[7]:
                            if values[3] == "CI":
                                if clin == "VS12110" or clin == "VS11310" or clin == "VS11210":
                                    if values[9] == "fixed":
                                        exempt = False
                                if exempt == False or values[9] == "percent":
                                    if clin[:2] == "EN" and values[8] == "E":
                                        valid = True
                                    if values[8] == "L":
                                        valid = True
                                    if values[8] == "B":
                                        valid = True
                                    if values[8] == "A":
                                        valid = True
                                    if values[8] == "S":
                                        valid = True
                                    if taxCodeCI == "" or values[8] != "B":
                                        if values[8] == "S" or values[8] == "A":
                                            valid = True
                                    if values[4] > highestModCI and valid == True:  # only take the highest tax mod
                                        highestModCI = values[4]
                                        taxCodeCI = key
                                        taxCodeCIAmt = values[5]
                                        taxCodeCIAmt = taxCode_dict[state][county][city][key][5]
                                        lineTaxTypeCI = taxCode_dict[state][county][city][key][9]
                        valid = True

        #  Just hard code DC to these values because they're weird
        taxCodeDC == ""
        if state == "DC":
            taxCodeST = "DC_ST4_153*1.1"
            taxCodeSTAmt = .11
            lineTaxTypeST = "percent"
            if clin == "VS12110" or clin == "VS11310" or clin == "VS11210":
                taxCodeDC = "DC_ST13_710"
                taxCodeDCAmt = 0.76
                lineTaxTypeDC = "fixed"

        # matches CLIN to federal tax(es)
        if clin in fed_dict:
            fed_values = fed_dict[clin]
            if len(fed_values) >= 4:
                federalTax1, federalTax2, federalTax3, federalTax4 = fed_values
                # check if fed tax exists and if it falls within date range
            for key in taxCode_dict["NA"]["NA"]["NA"]:
                if federalTax1 in key:  # IF EN_NS then no tax
                    if taxCode_dict["NA"]["NA"]["NA"][key][6] <= ubi_dateDict[ubi] <= \
                            taxCode_dict["NA"]["NA"]["NA"][key][7]:
                        federalTax1_Amt = taxCode_dict["NA"]["NA"]["NA"][key][5]
                        federalTax1_Type = taxCode_dict["NA"]["NA"]["NA"][key][9]
                # check if fed tax 1 exists and if it falls within date range
                if type(federalTax2) == str:
                    if federalTax2 in key:
                        if taxCode_dict["NA"]["NA"]["NA"][key][6] <= ubi_dateDict[ubi] <= \
                                taxCode_dict["NA"]["NA"]["NA"][key][7]:
                            federalTax2_Amt = taxCode_dict["NA"]["NA"]["NA"][key][5]
                            federalTax2_Type = taxCode_dict["NA"]["NA"]["NA"][key][9]

                # check if fed tax 1 exists and if it falls within date range
                if type(federalTax3) == str:
                    if federalTax3 in key:
                        if taxCode_dict["NA"]["NA"]["NA"][key][6] <= ubi_dateDict[ubi] <= \
                                taxCode_dict["NA"]["NA"]["NA"][key][7]:
                            federalTax3_Amt = taxCode_dict["NA"]["NA"]["NA"][key][5]
                            federalTax3_Type = taxCode_dict["NA"]["NA"]["NA"][key][9]

                # check if fed tax 1 exists and if it falls within date range
                if type(federalTax4) == str:
                    if federalTax4 in key:
                        if taxCode_dict["NA"]["NA"]["NA"][key][6] <= ubi_dateDict[ubi] <= \
                                taxCode_dict["NA"]["NA"]["NA"][key][7]:
                            federalTax4_Amt = taxCode_dict["NA"]["NA"]["NA"][key][5]
                            federalTax4_Type = taxCode_dict["NA"]["NA"]["NA"][key][9]

        # print to horizontal sheet
        # horizontal output needs to have total line tax calculated
        # city tax code calculation
        if taxCodeCI != "":
            output_sheet_horizontal["O" + str(currentRow_horizontal)] = remove_mod(taxCodeCI)
            if lineTaxTypeCI == "percent":
                lineTax = lineTax + (lineItemAmt * taxCodeCIAmt)
            if lineTaxTypeCI == "fixed":
                lineTax = lineTax + taxCodeCIAmt
            if remove_mod(taxCodeCI) == "AZ_MU4_17":
                output_sheet_horizontal["Q" + str(currentRow_horizontal)] = "AZ_MU3_2366"
                if lineTaxTypeCI == "percent":
                    lineTax = lineTax + (lineItemAmt * taxCodeCIAmt)
                if lineTaxTypeCI == "fixed":
                    lineTax = lineTax + taxCodeCIAmt

        # county tax code calculation
        if taxCodeCO != "":
            output_sheet_horizontal["N" + str(currentRow_horizontal)] = remove_mod(taxCodeCO)
            if lineTaxTypeCO == "percent":
                lineTax = lineTax + (lineItemAmt * taxCodeCOAmt)
            if lineTaxTypeCO == "fixed":
                lineTax = lineTax + taxCodeCOAmt
        if taxCodeST != "":
            output_sheet_horizontal["M" + str(currentRow_horizontal)] = remove_mod(taxCodeST)
            if lineTaxTypeST == "percent":
                lineTax = lineTax + (lineItemAmt * taxCodeSTAmt)
            if lineTaxTypeST == "fixed":
                lineTax = lineTax + taxCodeSTAmt

        # DC tax code calculation
        if taxCodeDC != "":
            output_sheet_horizontal["Q" + str(currentRow_horizontal)] = remove_mod(taxCodeDC)
            if lineTaxTypeST == "percent":
                lineTax = lineTax + (lineItemAmt * taxCodeSTAmt)
            if lineTaxTypeST == "fixed":
                lineTax = lineTax + taxCodeSTAmt

        # federal tax code calculations
        if federalTax1 != "":
            # federalTax1 is always the same if a federal tax exists
            output_sheet_horizontal["P" + str(currentRow_horizontal)] = federalTax1
            output_sheet_horizontal["A" + str(currentRow_horizontal)] = ubi_Whole
        lineTax = lineTax + (lineItemAmt * federalTax1_Amt)
        if federalTax2 != "" and type(federalTax2) == str:
            output_sheet_horizontal["V" + str(currentRow_horizontal)] = federalTax2
            if federalTax2_Type == "percent":
                lineTax = lineTax + (lineItemAmt * federalTax2_Amt)
            if federalTax2_Type == "fixed":
                lineTax = lineTax + federalTax2_Amt
        if federalTax3 != "" and type(federalTax3) == str:
            output_sheet_horizontal["W" + str(currentRow_horizontal)] = federalTax3
            if federalTax3_Type == "percent":
                lineTax = lineTax + (lineItemAmt * federalTax3_Amt)
            if federalTax3_Type == "fixed":
                lineTax = lineTax + federalTax3_Amt
        if federalTax4 != "" and type(federalTax4) == str:
            output_sheet_horizontal["X" + str(currentRow_horizontal)] = federalTax4
            if federalTax4_Type == "percent":
                lineTax = lineTax + (lineItemAmt * federalTax4_Amt)
            if federalTax4_Type == "fixed":
                lineTax = lineTax + federalTax4_Amt
        output_sheet_horizontal["Y" + str(currentRow_horizontal)] = lineTax
        output_sheet_horizontal["Z" + str(currentRow_horizontal)] = lineItemAmt
        currentRow_horizontal = currentRow_horizontal + 1

        # print taxes to vertical sheet
        # city tax code calculation
        if taxCodeCI != "":
            currentRow += 1
            toSheet(output_sheet, currentRow, ubi_Whole, taxCodeCI, lineItemAmt, dataTranCode, dataTranDate, dataTranLSN, ubi_valuesDict)
            taxCalc(lineTaxTypeCI, taxCodeCIAmt, lineItemAmt, proration, currentRow, output_sheet)

            if remove_mod(taxCodeCI) == "AZ_MU4_17":
                currentRow = currentRow + 1
                az_tax_code = "AZ_MU3_2366"
                toSheet(output_sheet, currentRow, ubi_Whole, az_tax_code, lineItemAmt, dataTranCode, dataTranDate,
                        dataTranLSN, ubi_valuesDict)
                output_sheet["N" + str(currentRow)] = (0.045 * lineItemAmt)

        # county tax code calculation
        if taxCodeCO != "":
            currentRow += 1
            toSheet(output_sheet, currentRow, ubi_Whole, taxCodeCO, lineItemAmt, dataTranCode, dataTranDate, dataTranLSN,
                    ubi_valuesDict)
            taxCalc(lineTaxTypeCO, taxCodeCOAmt, lineItemAmt, proration, currentRow, output_sheet)

        # state tax code calculation
        if taxCodeST != "":
            currentRow += 1
            toSheet(output_sheet, currentRow, ubi_Whole, taxCodeST, lineItemAmt, dataTranCode, dataTranDate,
                    dataTranLSN, ubi_valuesDict)
            taxCalc(lineTaxTypeST, taxCodeSTAmt, lineItemAmt, proration, currentRow, output_sheet)

        # DC tax code calculation
        if taxCodeDC != "":
            currentRow += 1
            toSheet(output_sheet, currentRow, ubi_Whole, taxCodeDC, lineItemAmt, dataTranCode, dataTranDate,
                    dataTranLSN, ubi_valuesDict)
            taxCalc(lineTaxTypeDC, taxCodeDCAmt, lineItemAmt, proration, currentRow, output_sheet)

        # federal tax code calculations
        if federalTax1 != "":
            currentRow = currentRow + 1
            toSheet(output_sheet, currentRow, ubi_Whole, federalTax1, lineItemAmt, dataTranCode, dataTranDate,
                    dataTranLSN, ubi_valuesDict)
            # federalTax1 is always the same if a federal tax exists
            output_sheet["N" + str(currentRow)] = federalTax1_Amt * lineItemAmt

        if federalTax2 != "" and type(federalTax2) == str:
            currentRow = currentRow + 1
            toSheet(output_sheet, currentRow, ubi_Whole, federalTax2, lineItemAmt, dataTranCode, dataTranDate,
                    dataTranLSN, ubi_valuesDict)
            taxCalc(federalTax2_Type, federalTax2_Amt, lineItemAmt, proration, currentRow, output_sheet)

        if federalTax3 != "" and type(federalTax3) == str:
            currentRow = currentRow + 1
            toSheet(output_sheet, currentRow, ubi_Whole, federalTax3, lineItemAmt, dataTranCode, dataTranDate,
                    dataTranLSN, ubi_valuesDict)
            taxCalc(federalTax3_Type, federalTax3_Amt, lineItemAmt, proration, currentRow, output_sheet)

        if federalTax4 != "" and type(federalTax4) == str:
            currentRow = currentRow + 1
            toSheet(output_sheet, currentRow, ubi_Whole, federalTax4, lineItemAmt, dataTranCode, dataTranDate,
                    dataTranLSN, ubi_valuesDict)
            taxCalc(federalTax4_Type, federalTax4_Amt, lineItemAmt, proration, currentRow, output_sheet)

        # loop operations
        currentRead = currentRead + 1
        print("Currently on row: " + str(currentRead - 1) + " of " + str(len(ubi_list)))
        ubi_line_count = ubi_line_count + 1
        nsc_line_count = nsc_line_count + 1
except Exception as e:
    print("Bad spreadsheet data. Columns were likely adjusted from original format or database error.")
    input("Press Enter to exit...")

# avoids annoyance of having output sheet open upon completion
try:
    ubi_output_horizontal.save("Horizontal" + sheetName + "_" + excelName)
    ubi_output.save(sheetName + "_" + excelName)
except Exception as e:
    print("Problem with saving output. Ensure relevant files are not already opened")
    input("Press Enter to exit...")

print("")
print('Complete')

input("Press any Enter to exit...")
